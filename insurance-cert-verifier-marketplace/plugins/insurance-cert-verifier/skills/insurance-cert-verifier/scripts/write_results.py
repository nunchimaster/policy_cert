"""검증 결과를 엑셀에 기록 — **원본 서식을 보존**하고 해당 셀만 갱신한다.

핵심: 결과 파일은 원본 양식(열너비/병합/폰트/색)을 그대로 유지하고
'대상요약'에 판정 컬럼만 덧쓰고, '검증상세' 시트를 추가한다.

- 입력이 .xlsx  : openpyxl로 원본을 **열어** 해당 셀만 덮어쓴다(서식 보존).
- 입력이 .xls   : 먼저 .xlsx로 변환(LibreOffice `soffice` 헤드리스; 없으면 Windows의 Excel COM)
                  후 위와 동일. 변환 수단이 없으면 값만 복제하는 레거시 방식으로 폴백(서식 손실, 경고).

대상요약 갱신:
    I열(9)=확인내용, M~Q(13~17)=검증PDF/검증페이지/텍스트판정/금액판정/종합사유
    매칭은 보종코드(B열=2) 기준. 헤더행은 4행(1-index).
검증상세(신규):
    보종코드/보종명/증권번호/PDF파일/페이지/항목구분/엑셀값/증권출력값/결과/사유
    (보장명칭 H, 보장내역 I, 보장금액 라인이 각각 1행)

사용법:
  python write_results.py --xls 원본.(xls|xlsx) --results results.json --out 원본_검증.xlsx
"""
import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

RESULT_FILL = {
    "O":  PatternFill("solid", fgColor="C6EFCE"),
    "X":  PatternFill("solid", fgColor="FFC7CE"),
    "검토": PatternFill("solid", fgColor="FFEB9C"),
}
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
HDR_FONT = Font(bold=True)


def code_str(v):
    if isinstance(v, float):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def _convert_xls_to_xlsx(xls_path):
    """.xls -> .xlsx 변환 (서식 보존). 성공 시 xlsx 경로, 실패 시 None."""
    tmpdir = tempfile.mkdtemp(prefix="certverify_")
    # 1) LibreOffice (리눅스/맥/윈도우 공통, 헤드리스)
    for exe in ("soffice", "libreoffice"):
        if shutil.which(exe):
            try:
                subprocess.run([exe, "--headless", "--convert-to", "xlsx",
                                "--outdir", tmpdir, xls_path],
                               check=True, capture_output=True, timeout=120)
                cand = os.path.join(tmpdir, os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx")
                if os.path.exists(cand):
                    return cand
            except Exception:
                pass
    # 2) Windows Excel COM
    if sys.platform.startswith("win"):
        try:
            import win32com.client  # type: ignore
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(xls_path))
            cand = os.path.join(tmpdir, os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx")
            wb.SaveAs(os.path.abspath(cand), FileFormat=51)  # 51 = xlOpenXMLWorkbook
            wb.Close(False)
            excel.Quit()
            if os.path.exists(cand):
                return cand
        except Exception:
            pass
    return None


def _legacy_rebuild(xls_path, results, out):
    """변환 수단이 없을 때만: 값만 복제(서식 손실). 경고."""
    import xlrd
    print("[경고] .xls 변환 수단(soffice/Excel)이 없어 값만 복제합니다(원본 서식 손실).")
    wb_in = xlrd.open_workbook(xls_path)
    wb = Workbook(); wb.remove(wb.active)
    sheet_ws = {}
    for name in wb_in.sheet_names():
        ws = wb.create_sheet(title=name[:31])
        s = wb_in.sheet_by_name(name)
        for r in range(s.nrows):
            for c in range(s.ncols):
                v = s.cell_value(r, c)
                if v != "":
                    ws.cell(row=r + 1, column=c + 1, value=v)
        sheet_ws[name] = ws
    _update_summary_and_detail(wb, sheet_ws["대상요약"], results)
    wb.save(out)


def _update_summary_and_detail(wb, ws1, results):
    by_code = {r["보종코드"]: r for r in results}
    HDR = 4
    summary_cols = [(13, "검증PDF"), (14, "검증페이지"), (15, "텍스트판정"),
                    (16, "금액판정"), (17, "종합사유")]
    for c, title in summary_cols:
        cell = ws1.cell(row=HDR, column=c, value=title)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
    for r in range(HDR + 1, ws1.max_row + 1):
        code = code_str(ws1.cell(row=r, column=2).value)
        res = by_code.get(code)
        if not res:
            continue
        ws1.cell(row=r, column=9, value=res.get("확인내용", ""))
        ws1.cell(row=r, column=13, value=res.get("검증PDF", ""))
        ws1.cell(row=r, column=14, value=res.get("검증페이지", ""))
        for c, key in [(15, "텍스트판정"), (16, "금액판정")]:
            v = res.get(key, "")
            cell = ws1.cell(row=r, column=c, value=v)
            if v in RESULT_FILL:
                cell.fill = RESULT_FILL[v]
        ws1.cell(row=r, column=17, value=res.get("종합사유", ""))

    if "검증상세" in wb.sheetnames:
        del wb["검증상세"]
    ws2 = wb.create_sheet(title="검증상세")
    headers = ["보종코드", "보종명", "증권번호", "PDF파일", "페이지",
               "항목구분", "엑셀값", "증권출력값", "결과", "사유"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
    for c, wdt in enumerate([10, 34, 13, 30, 8, 26, 46, 22, 6, 40], 1):
        ws2.column_dimensions[chr(64 + c)].width = wdt
    row = 2
    for res in results:
        for ln in res.get("lines", []):
            vals = [res.get("보종코드"), res.get("보종명"), res.get("증권번호"),
                    res.get("검증PDF"), ln.get("페이지"),
                    ln.get("항목구분"), ln.get("엑셀값"), ln.get("증권출력값"),
                    ln.get("결과"), ln.get("사유")]
            for c, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=c, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            rcell = ws2.cell(row=row, column=9)
            if ln.get("결과") in RESULT_FILL:
                rcell.fill = RESULT_FILL[ln["결과"]]
            row += 1
    return row - 2


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xls", required=True)
    ap.add_argument("--results", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    results = json.load(open(args.results, encoding="utf-8"))
    src = args.xls
    ext = os.path.splitext(src)[1].lower()

    if ext == ".xlsx":
        wb = load_workbook(src)            # 원본 서식 그대로 로드
    else:                                   # .xls -> 변환 시도
        conv = _convert_xls_to_xlsx(src)
        if conv:
            wb = load_workbook(conv)
        else:
            _legacy_rebuild(src, results, args.out)
            print(f"기록 완료(레거시): {args.out}")
            return

    if "대상요약" not in wb.sheetnames:
        raise SystemExit("'대상요약' 시트를 찾을 수 없습니다.")
    n = _update_summary_and_detail(wb, wb["대상요약"], results)
    wb.save(args.out)
    print(f"기록 완료(서식 보존): {args.out}  (검증상세 {n}행)")


if __name__ == "__main__":
    main()
