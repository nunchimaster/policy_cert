"""검증용 엑셀(.xls/.xlsx) 파싱.

- '대상요약'      : 보종코드/보종명/증권번호(='테스트 케이스')/담당자/확인내용
                    + 변형 플래그(S기본/S무해지/W기본/W무해지 등 'O' 표시) 자동 인식.
- 상세 시트       : **헤더 시그니처로 자동 선별**한다. 상세시트는 3번째 행(0-index 2)의
                    A열이 'PLICD'인 시트(보종코드/상품명/.../보장내용01/02/수식01/분자/분모...).
                    'W(57)'·'S(314)' 같은 증권 구성표(상세 아님)는 제외된다.
  컬럼: 보종코드(A,0) / 라인01(D,3) / 보장내용01=H(7) / 보장내용02=I(8)
        / 수식01(J,9) / 분자01(L,11) / 분모01(M,12).  (금액은 수식01만 사용)

여러 상세 시트가 있으면 코드별로 누적 병합한다.

사용법:
  python parse_excel.py --xls 원본.(xls|xlsx) --out parsed.json [--code 86187]
"""
import argparse
import json
import os
from verify_lib import to_num

DETAIL_ROW0 = 3        # 상세 데이터 시작행(1-index 4 == 0-index 3)
DETAIL_SIG_ROW = 2     # 0-index 2행 A열 == 'PLICD' 이면 상세시트
MAXCOL = 13            # 필요한 최대 열(M=12) + 1


def code_str(v):
    if isinstance(v, float):
        return str(int(v))
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def load_grids(path):
    """엑셀을 {시트명: 2D리스트(최대 MAXCOL열)} 로. .xls=xlrd, .xlsx=openpyxl."""
    ext = os.path.splitext(path)[1].lower()
    order, sheets = [], {}
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        order = list(wb.sheetnames)
        for name in order:
            ws = wb[name]
            sheets[name] = [list(r) for r in ws.iter_rows(max_col=MAXCOL, values_only=True)]
        wb.close()
    else:
        import xlrd
        wb = xlrd.open_workbook(path)
        order = wb.sheet_names()
        for name in order:
            s = wb.sheet_by_name(name)
            sheets[name] = [[(s.cell_value(r, c) if c < s.ncols else None)
                             for c in range(MAXCOL)] for r in range(s.nrows)]
    return order, sheets


def g(grid, r, c):
    if 0 <= r < len(grid) and 0 <= c < len(grid[r]):
        v = grid[r][c]
        return "" if v is None else v
    return ""


def is_detail(grid):
    return str(g(grid, DETAIL_SIG_ROW, 0)).strip().upper() == "PLICD"


def parse_summary(grid):
    """대상요약: 헤더행을 찾아 컬럼명→인덱스 매핑 후 행을 읽는다."""
    # 헤더행 = '보종코드'가 있는 행
    hdr = None
    for r in range(min(12, len(grid))):
        rowvals = [str(g(grid, r, c)).strip() for c in range(MAXCOL)]
        if "보종코드" in rowvals:
            hdr = r
            colmap = {rowvals[c]: c for c in range(MAXCOL) if rowvals[c]}
            break
    if hdr is None:
        # 기존 고정 레이아웃 폴백 (B=1 코드, C=2 명, G=6 증권, H=7 담당)
        hdr, colmap = 3, {"보종코드": 1, "보종명": 2, "테스트 케이스": 6, "담당자": 7}

    ci = lambda *names: next((colmap[n] for n in names if n in colmap), None)
    c_code = ci("보종코드"); c_name = ci("보종명")
    c_cert = ci("테스트 케이스", "증권번호"); c_mgr = ci("담당자")
    c_chk = ci("확인내용")
    # 변형 플래그 컬럼(S기본/S무해지/W기본/W무해지 등) — 헤더에 그대로 있으면 캡처
    flag_cols = {n: i for n, i in colmap.items()
                 if n in ("S기본", "S무해지", "W기본", "W무해지") or n.endswith(("기본", "무해지"))}

    targets = []
    for r in range(hdr + 1, len(grid)):
        code = code_str(g(grid, r, c_code)) if c_code is not None else ""
        if not code or not code.replace(".", "").isdigit():
            continue
        cert_raw = g(grid, r, c_cert) if c_cert is not None else ""
        cert = code_str(cert_raw) if isinstance(cert_raw, float) else str(cert_raw).strip()
        flags = [n for n, i in flag_cols.items() if str(g(grid, r, i)).strip()]
        targets.append({
            "보종코드": code,
            "보종명": str(g(grid, r, c_name)).strip() if c_name is not None else "",
            "증권번호": cert,
            "담당자": str(g(grid, r, c_mgr)).strip() if c_mgr is not None else "",
            "확인내용현재": str(g(grid, r, c_chk)).strip() if c_chk is not None else "",
            "변형": flags,          # 예: ['S기본'] — 어느 PDF에 있는지
            "xls_row": r,
        })
    return targets


def parse(xls_path):
    order, sheets = load_grids(xls_path)
    summary_name = "대상요약" if "대상요약" in sheets else next(
        (n for n in order if "대상요약" in n), order[0])
    targets = parse_summary(sheets[summary_name])

    detail_names = [n for n in order if n != summary_name and is_detail(sheets[n])]
    coverage = {}
    for name in detail_names:
        grid = sheets[name]
        for r in range(DETAIL_ROW0, len(grid)):
            code = code_str(g(grid, r, 0))
            if not code or not code.isdigit():
                continue
            H = str(g(grid, r, 7)).strip()
            I = str(g(grid, r, 8)).strip()
            수식 = str(g(grid, r, 9)).strip()
            분자 = to_num(g(grid, r, 11))
            분모 = to_num(g(grid, r, 12))
            라인 = g(grid, r, 3)
            cov = coverage.setdefault(code, {
                "상품명": str(g(grid, r, 1)).strip(), "시트": name,
                "명칭라인": [], "내역라인": [], "금액라인": [],
            })
            if H:
                cov["명칭라인"].append(H)
            if I:
                cov["내역라인"].append(I)
            if 분자 > 0 and 분모 > 0:
                cov["금액라인"].append({"텍스트": I, "수식구분코드": 수식,
                                     "분자": int(분자), "분모": int(분모), "라인": str(라인)})

    for cov in coverage.values():
        cov["보장명칭"] = "".join(cov["명칭라인"])
        cov["보장내역"] = "\n".join(cov["내역라인"])

    return {"summary_sheet": summary_name, "detail_sheets": detail_names,
            "대상요약": targets, "보장": coverage}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xls", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--code")
    args = ap.parse_args()
    data = parse(args.xls)
    if args.code:
        data["대상요약"] = [t for t in data["대상요약"] if t["보종코드"] == args.code]
        data["보장"] = {k: v for k, v in data["보장"].items() if k == args.code}
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[{data['summary_sheet']}] 대상요약 {len(data['대상요약'])}행 / "
          f"상세시트 {len(data['detail_sheets'])}개 / 보장 {len(data['보장'])}코드 -> {args.out}")
    print("detail_sheets:", data["detail_sheets"])


if __name__ == "__main__":
    main()
